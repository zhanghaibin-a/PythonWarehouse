#coding=utf-8
from selenium import webdriver
from lxml import etree
import openpyxl
import time
option = webdriver.ChromeOptions()
option.add_argument('headless')

class lazada(object):
    def __init__(self,url=None):
        self.home_url=url
        self.lazada_list=[]
    def response_url(self):
        #获取页面所有宝贝的url
        #1.请求页面
        Chrome_web = webdriver.Chrome(executable_path='chromedriver', chrome_options=option)
        Chrome_web.get(self.home_url)
        time.sleep(25)
        #2.滚动网页
        for y in range(30):
            js = 'window.scrollBy(0,150)'
            Chrome_web.execute_script(js)
            time.sleep(1)
        #3.过滤url
        xpath_data = etree.HTML(Chrome_web.page_source)
        url_list = xpath_data.xpath("//div[@class='cRjKsc']/a/@href")
        Chrome_web.quit()
        #4.返回数据list
        return url_list
        pass
    def url_info(self,url_list):
        #迭代请求
        for i in url_list:
            str_url='https://'+i.replace('//','')
            Chrome_web = webdriver.Chrome(executable_path='chromedriver', chrome_options=option)#executable_path='chromedriver', chrome_options=option
            Chrome_web.get(str_url)
            time.sleep(25)
            #滚动网页
            for y in range(5):
                js = 'window.scrollBy(0,100)'
                Chrome_web.execute_script(js)
                time.sleep(3)
            #过滤
            try:
                xpath_data = etree.HTML(Chrome_web.page_source)
                headline = xpath_data.xpath("//span[@class='pdp-mod-product-badge-title']/text()")  # headline标题
                shop_name = xpath_data.xpath("//a[@class='pdp-link pdp-link_size_l pdp-link_theme_black seller-name__detail-name']/text()")#shop_name店铺名
                try:
                    Penilaian_Pertanyaan_yang_dijawab = xpath_data.xpath(
                        "//a[@class='pdp-link pdp-link_size_s pdp-link_theme_blue pdp-review-summary__link']/text()")  # Pertanyaan_yang_dijawab, Penilaian
                    Penilaian_Pertanyaan_yang_dijawab.extend ([1,2])
                    Penilaian_positif_seller = xpath_data.xpath(
                        '//div[@class="seller-info-value rating-positive "]/text()')  # Penilaian_positif_seller
                    PPengiriman_tepat_waktu,Tingkat_respon_chat = xpath_data.xpath(
                        '//div[@class="seller-info-value "]/text()')  # PPengiriman_tepat_waktu,Tingkat_respon_chat
                except ValueError :
                    print("过滤:")
            except ValueError as err :
                print("过滤:"+err)
            #生成字典格式
            try:
                lazada_dict={}
                lazada_dict['url']=str_url
                lazada_dict['headline']=headline
                lazada_dict['shop_name'] =shop_name
                lazada_dict['Penilaian'] =Penilaian_Pertanyaan_yang_dijawab[0]
                lazada_dict['Pertanyaan_yang_dijawab'] =Penilaian_Pertanyaan_yang_dijawab[1] #Pertanyaan_yang_dijawab
                lazada_dict['Penilaian_positif_seller'] =Penilaian_positif_seller
                lazada_dict['PPengiriman_tepat_waktu'] =PPengiriman_tepat_waktu
                lazada_dict['Tingkat_respon_chat'] = Tingkat_respon_chat
                Chrome_web.quit()
            except BaseException as err:
                print("字典:" + err)
            #获取详情页的相关信息添加到List
            self.lazada_list.append(lazada_dict)
            print(lazada_dict)
    def save_excel(self,save_data,file_name='test_case2.xlsx'):
        try:
            """保存信息"""
            excel = openpyxl.load_workbook(file_name)
            excel_sheet = excel['Sheet1']
            nub = 1
            for data_info in save_data:
                nub = nub + 1
                excel_sheet.cell(row=nub, column=1, value=str(str(data_info['url'])))
                excel_sheet.cell(row=nub, column=2, value=str(str(data_info['headline'][0])))
                excel_sheet.cell(row=nub, column=3, value=str(str(data_info['shop_name'][0])))
                excel_sheet.cell(row=nub, column=4, value=str(str(data_info['Penilaian'])))
                excel_sheet.cell(row=nub, column=5, value=str(str(data_info['Pertanyaan_yang_dijawab'])))
                excel_sheet.cell(row=nub, column=6, value=str(str(data_info['Penilaian_positif_seller'])))
                excel_sheet.cell(row=nub, column=7, value=str(str(data_info['PPengiriman_tepat_waktu'])))
                excel_sheet.cell(row=nub, column=8, value=str(str(data_info['Tingkat_respon_chat'])))
            excel.save(file_name)
            excel.close()
        except BaseException as err:
            print("save_excel:" + err)
requests=lazada(url='https://www.lazada.co.id/catalog/?q=baju+pria&_keyori=ss&from=input&spm=a2o4j.searchlist.search.go.69f132d0XMcAHa')#url='https://www.lazada.co.id/products/blcod-baju-kaos-distrolengan-panjangdanilogood-quality-i460422163-s656724135.html?search=1'
url_list=requests.response_url()
print(len(url_list))
print(url_list)
requests.url_info(url_list=url_list)
data=requests.lazada_list
print(data)
#requests.save_excel(save_data=data)