#coding=utf-8
from selenium import webdriver
from lxml import etree
import openpyxl
import time
option = webdriver.ChromeOptions()
option.add_argument('headless')

class shoppe(object):
    def __init__(self,url=None):
        self.home_url=url
        self.shoppe_list=[]
    def response_url(self):
        #获取页面所有宝贝的url
        #1.请求页面
        Chrome_web = webdriver.Chrome(executable_path='chromedriver', chrome_options=option)
        Chrome_web.get(self.home_url)
        time.sleep(25)
        #2.滚动网页
        for y in range(30):
            js = 'window.scrollBy(0,100)'
            Chrome_web.execute_script(js)
            time.sleep(1)
        #3.过滤url
        xpath_data = etree.HTML(Chrome_web.page_source)
        url_list = xpath_data.xpath("//a[@data-sqe]//@href")
        Chrome_web.quit()
        #4.返回数据list
        return url_list
        pass
    def url_info(self,url_list):
        #迭代请求
        for i in url_list:
            str_url="https://shopee.co.id"+i
            Chrome_web = webdriver.Chrome(executable_path='chromedriver', chrome_options=option)
            Chrome_web.get(str_url)
            time.sleep(25)
            #滚动网页
            for y in range(20):
                js = 'window.scrollBy(0,100)'
                Chrome_web.execute_script(js)
                time.sleep(3)
            #过滤
            try:
                xpath_data = etree.HTML(Chrome_web.page_source)
                headline = xpath_data.xpath('//div[@class="qaNIZv"]/span/text()')  # 标题
                shop_name = xpath_data.xpath('//div[@class="_3Lybjn"]/text()')  # 店铺名
                try:
                    Penilaian,Produk,Persentase = xpath_data.xpath('//span[@class="_1rsHot OuQDPE"]/text()')  # Penilaian，Produk，Persentase Chat Dibalas
                    Waktu_Chat_Dibalas,Bergabung,Pengikut = xpath_data.xpath('//span[@class="_1rsHot"]/text()')  # Waktu Chat Dibalas，Bergabung，Pengikut
                except ValueError:
                    print("过滤:")
            except ValueError as err :
                print("过滤:"+err)
            #生成字典格式
            try:
                shoppe_dict={}
                shoppe_dict['url']=str_url
                shoppe_dict['headline']=headline
                shoppe_dict['headline'] =headline
                shoppe_dict['shop_name'] =shop_name
                shoppe_dict['Penilaian'] =Penilaian
                shoppe_dict['Produk'] =Produk
                shoppe_dict['Persentase'] =Persentase
                shoppe_dict['Waktu_Chat_Dibalas'] =Waktu_Chat_Dibalas
                shoppe_dict['Bergabung'] =Bergabung
                shoppe_dict['Pengikut'] =Pengikut
                Chrome_web.quit()
            except BaseException as err:
                print("字典:" + err)
            #获取详情页的相关信息添加到List
            self.shoppe_list.append(shoppe_dict)
            print(shoppe_dict)
    def save_excel(self,save_data,file_name='test_case.xlsx'):
        try:
            """保存信息"""
            excel = openpyxl.load_workbook(file_name)
            excel_sheet = excel['Sheet1']
            nub = 1
            for data_info in save_data:
                nub = nub + 1
                excel_sheet.cell(row=nub, column=1, value=str(str(data_info['url'])))
                excel_sheet.cell(row=nub, column=2, value=str(str(data_info['headline'][0])))
                excel_sheet.cell(row=nub, column=3, value=str(str(data_info['shop_name'][0])))
                excel_sheet.cell(row=nub, column=4, value=str(str(data_info['Penilaian'])))
                excel_sheet.cell(row=nub, column=5, value=str(str(data_info['Produk'])))
                excel_sheet.cell(row=nub, column=6, value=str(str(data_info['Persentase'])))
                excel_sheet.cell(row=nub, column=7, value=str(str(data_info['Waktu_Chat_Dibalas'])))
                excel_sheet.cell(row=nub, column=8, value=str(str(data_info['Bergabung'])))
                excel_sheet.cell(row=nub, column=9, value=str(str(data_info['Pengikut'])))
            excel.save(file_name)
            excel.close()
        except BaseException as err:
            print("save_excel:" + err)
requests=shoppe(url='https://shopee.co.id/search?keyword=e-cigarette')
url_list=requests.response_url()
print(len(url_list))
print(url_list)
requests.url_info(url_list=url_list)
data=requests.shoppe_list
print(data)
requests.save_excel(save_data=data)