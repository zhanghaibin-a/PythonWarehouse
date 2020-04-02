#coding=utf-8

from ldap3 import Server, Connection, ALL, MODIFY_REPLACE,AUTO_BIND_NO_TLS
from datetime import  timedelta, datetime
import openpyxl
"""
输入LDAP服务器相关信息
"""
server = Server('ldaps://192.168.26.129',get_info=ALL,use_ssl=True)
admin_username = 'aabbcc\\Administrator'
admin_password = 'zhangHAIbin0.'
group_dn = 'OU=test,DC=aabbcc,DC=com' #OU为组织单位,DC为域名拆分

def add_ad_user(add_username,name,password, phone, mail,user_overdue_day=30):
    """添加域账号"""
    user_dn ='cn=%s,%s' %(add_username,group_dn)
    try:
        """创建初始化连接"""
        c = Connection(server, auto_bind=True, user=admin_username, password=admin_password)
        """获取当前系统时间"""
        end_time = datetime.today() + timedelta(days=user_overdue_day)
        """添加用户"""
        c.add(user_dn, attributes={'objectClass':  ['top', 'person', 'organizationalPerson', 'user'],
                                   # 用户名
                                   'sAMAccountName': add_username,
                                   # 用户名
                                   'userPrincipalName': add_username,
                                   # 有效期几天
                                   'accountExpires': end_time,
                                   # name可为中文
                                   'sn': name,
                                   # 显示名为用户名
                                   'displayName': add_username,
                                   # 电话
                                   "telephoneNumber": phone,
                                   # 邮件
                                   "Mail": mail,
                                   })
        # 添加用户到组
        print(end_time)
        c.extend.microsoft.add_members_to_groups(user_dn, group_dn)
        c.extend.microsoft.modify_password(user=user_dn,new_password=password)
        # 激活用户
        print(user_dn)
        c.modify(user_dn,{'userAccountControl': [(MODIFY_REPLACE, [544,])]})#账户正常使用
        return end_time,user_dn

    except Exception as err:
        print(err)
        return None

def increase_of_service_life(user_dn,days=180):
    """修改密码有效期"""
    try:
        c = Connection(server, auto_bind=AUTO_BIND_NO_TLS, read_only=False, check_names=True, user=admin_username, password=admin_password)
        c.modify(user_dn,
                 {'accountExpires': [(MODIFY_REPLACE, [datetime.now() + timedelta(days=days)])]})
    except Exception as e:
        print(e)
        return None

def expiration_time(user_dn):
    """获取用户过期时间"""
    try:
        # 连接服务器
        c = Connection(server, auto_bind=AUTO_BIND_NO_TLS, read_only=True, check_names=True, user=admin_username, password=admin_password)
        c.search(search_base=user_dn,
                 search_filter='(objectCategory=user)',
                 attributes=['accountExpires'],
                 size_limit=1)
        """初始化返回用户过期时间"""
        time_user=str(c.entries[0]).split(" - ")[-1].split("\n")[1].split(":")[1].replace(' ','-')
        return time_user
    except Exception as e:
        print(e)
        return None


class read_excel_info():
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name

    def read_excel(self):
        """读取excel相关信息,进行账号创建"""
        excel = openpyxl.load_workbook(self.file_name)
        excel_sheet = excel[self.sheet_name]
        nub = 1
        for case in list(excel_sheet.rows)[1:]:
            nub = nub+1
            acse=[]
            for i in range(len(case)):
                acse.append(case[i].value)
            try:
                if acse[0]==None or acse[1]==None:
                    """excel表中当行中包含格式也会被算作一行,用来过滤None行"""
                    print("当前行为None，自动跳转下一行")
                elif acse[6]==None or acse[7]==None:
                    """当 user_cn 或 Password_expiration_Time 为空时添加用户"""
                    print(acse)
                    end_time,user_dn=add_ad_user(add_username=acse[0],name=acse[1],password=acse[2],phone=acse[3],mail=acse[4],user_overdue_day=int(acse[5]))
                    excel_sheet.cell(row=nub,column=7,value =str(end_time))
                    excel_sheet.cell(row=nub, column=8, value=str(user_dn))
                elif acse[8] != None and acse[7] != None :
                    """当extend_passwd_time和 user_cn不为空时为账户续期"""
                    increase_of_service_life(user_dn=acse[7],days=int(acse[8]))
                    user_time=str(expiration_time(user_dn=acse[7]))
                    print("账户：%s，延长使用至：%s"%(acse[0],user_time))
                    excel_sheet.cell(row=nub, column=7,value=str(user_time))
                    excel_sheet.cell(row=nub, column=9, value='')
                else: pass
            except Exception as err:
                print("读取excel发送错误：%s"%err)
        excel.save('test_case.xlsx')
        excel.close()

if __name__ == '__main__':
    # for i in range(1, 5):
    #    print(add_ad_user('zhangsan%s'% str(i), '张三', 'GKndd14@bd', '1101010', '2350512555@qq.com'))
    #print(expiration_time("cn=zhangsan2,ou=aabbcc,DC=aabbcc,DC=com"))
    excel=read_excel_info("test_case.xlsx","Sheet1")
    excel.read_excel()